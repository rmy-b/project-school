from django.shortcuts import render,redirect
from django.contrib.auth.decorators import login_required
from details.models import Faculty, FacultyAssignment,Student, Marks, Exam, Attendance,Class
import datetime
from django.db.models import Avg
from faculty.ML.predictor import predict_pass_percentage
from faculty.ML.faculty_cluster import cluster_students


@login_required
def faculty_dashboard(request):
    user = request.user   # CustomUser

    # Get Faculty linked to this user
    faculty = Faculty.objects.get(user=user)

    # Get all assignments of this faculty
    assignments = FacultyAssignment.objects.filter(faculty=faculty)

    # Incharge class (if any)
    incharge = assignments.filter(is_class_incharge=True).first()

    # Classes handled (unique)
    classes_handled = assignments.values(
        'class_obj__class_name',
        'section__section_name',
        'subject__subject_name'
    ).distinct()

    context = {
        "faculty": faculty,
        "incharge": incharge,
        "classes_handled": classes_handled,
        "is_class_incharge": bool(incharge),
    }

    return render(request, "faculty/dashboard.html", context)

@login_required
def marks_page(request):
    faculty = Faculty.objects.get(user=request.user)

    is_class_incharge = FacultyAssignment.objects.filter(
    faculty=faculty,
    is_class_incharge=True
).exists()
    # All classes this teacher handles
    assignments = FacultyAssignment.objects.filter(faculty=faculty)
    exams = Exam.objects.all()

    selected_class_id = request.GET.get('class')
    selected_exam_id = request.GET.get('exam')

    students = []
    marks_data = []

    # Default class = class incharge class
    default_assignment = assignments.filter(is_class_incharge=True).first()

    if not selected_class_id and default_assignment:
        selected_class_id = str(default_assignment.class_obj.id)

    selected_assignment = None

    if selected_class_id:
        selected_assignment = assignments.filter(class_obj_id=selected_class_id).first()

    # ---------------- SAVE MARKS (POST) ----------------
    if request.method == "POST":
        selected_class_id = request.POST.get("class")
        selected_exam_id = request.POST.get("exam")

        selected_assignment = assignments.filter(class_obj_id=selected_class_id).first()
        exam = Exam.objects.filter(id=selected_exam_id).first()

        # Safety check (prevents crash)
        if not exam or not selected_assignment:
            return redirect(f"/faculty/marks/?class={selected_class_id}&exam={selected_exam_id}")

        students = Student.objects.filter(
            class_obj=selected_assignment.class_obj,
            section=selected_assignment.section
        )

        for student in students:
            internal = request.POST.get(f"internal_{student.id}")
            external = request.POST.get(f"external_{student.id}")

            if internal is not None and external is not None:
                internal_val = float(internal or 0)
                external_val = float(external or 0)
                total = internal_val + external_val

                Marks.objects.update_or_create(
                    student=student,
                    subject=selected_assignment.subject,
                    exam=exam,
                    defaults={
                        "internal_marks": internal_val,
                        "external_marks": external_val,
                        "total_marks": total,
                    }
                )

        # Redirect after POST (prevents resubmission)
        return redirect(f"/faculty/marks/?class={selected_class_id}&exam={selected_exam_id}")
    # ---------------------------------------------------

    # ---------------- DISPLAY MARKS (GET) ----------------
    if selected_assignment and selected_exam_id:
        exam = Exam.objects.filter(id=selected_exam_id).first()

        if exam:
            students = Student.objects.filter(
                class_obj=selected_assignment.class_obj,
                section=selected_assignment.section
            )

            for student in students:
                mark = Marks.objects.filter(
                    student=student,
                    subject=selected_assignment.subject,
                    exam=exam
                ).first()

                marks_data.append({
                    'student': student,
                    'internal': mark.internal_marks if mark else '',
                    'external': mark.external_marks if mark else '',
                    'total': mark.total_marks if mark else '',
                })
    # -----------------------------------------------------

    context = {
        'faculty': faculty,
        'assignments': assignments,
        'students': students,
        'marks_data': marks_data,
        'exams': exams,
        'selected_exam_id': selected_exam_id,
        'selected_class_id': selected_class_id,
        'selected_assignment': selected_assignment,
         'is_class_incharge': is_class_incharge,
    }

    return render(request, 'faculty/marks.html', context)

@login_required
def attendance_page(request):
    faculty = Faculty.objects.get(user=request.user)
    is_class_incharge = FacultyAssignment.objects.filter(
     faculty=faculty,
     is_class_incharge=True
).exists()
    # Check if this faculty is class incharge
    class_incharge = FacultyAssignment.objects.filter(
        faculty=faculty,
        is_class_incharge=True
    ).first()

    # Not class teacher
    if not class_incharge:
        return render(request, 'faculty/attendance.html', {
         'faculty': faculty,
        'not_authorized': True,
        'is_class_incharge': is_class_incharge,
    })

    # Class teacher
    class_obj = class_incharge.class_obj
    section = class_incharge.section

    students = Student.objects.filter(
        class_obj=class_obj,
        section=section
    )

    error_message = None

    # Get date
    if request.method == "POST":
        selected_date = request.POST.get("date")

        if not selected_date:
            error_message = "Please select a date"
            selected_date = datetime.date.today()
        else:
            selected_date = datetime.datetime.strptime(selected_date, "%Y-%m-%d").date()
    else:
        selected_date = datetime.date.today()

    # ---------------- SAVE ATTENDANCE ----------------
    if request.method == "POST" and not error_message:
        for student in students:
            checkbox_value = request.POST.get(f"present_{student.id}")

            status = "P" if checkbox_value == "on" else "A"

            Attendance.objects.update_or_create(
                student=student,
                date=selected_date,
                defaults={
                    'status': status
                }
            )
    # ------------------------------------------------

    # Load existing attendance
    attendance_data = []
    present_count = 0
    absent_count = 0

    for student in students:
        record = Attendance.objects.filter(
            student=student,
            date=selected_date
        ).first()

        if record:
            status = record.status
        else:
            status = "P"  # default

        if status == "P":
            present_count += 1
        else:
            absent_count += 1

        total_days = Attendance.objects.filter(student=student).count()
        present_days = Attendance.objects.filter(student=student, status="P").count()

        if total_days > 0:
            percentage = round((present_days / total_days) * 100)
        else:
            percentage = 0


        attendance_data.append({
            'student': student,
            'status': status,
            'percentage': percentage  
        })

    context = {
        'faculty': faculty,
        'class_obj': class_obj,
        'section': section,
        'students': students,
        'attendance_data': attendance_data,
        'selected_date': selected_date,
        'present_count': present_count,
        'absent_count': absent_count,
        'total_count': students.count(),
        'error_message': error_message,
        'not_authorized': False,
        'is_class_incharge':is_class_incharge,
    }

    return render(request, 'faculty/attendance.html', context)




def performance_view(request):
    faculty = Faculty.objects.get(user=request.user)
    is_class_incharge = FacultyAssignment.objects.filter(
    faculty=faculty,
    is_class_incharge=True
).exists()
    latest_exam = Exam.objects.order_by('-id').first()
    # Only classes assigned to this faculty
    assignments = FacultyAssignment.objects.filter(faculty=faculty)

    classes = assignments.values(
        'class_obj__id',
        'class_obj__class_name',
        'section__section_name'
    ).distinct()

    selected_class_id = request.GET.get("class")

    data = {
         "faculty": faculty,
         "classes": classes,
         "selected_class_id": selected_class_id,
         "is_class_incharge": is_class_incharge,
}

    if selected_class_id:
        faculty_subjects = assignments.filter(
            class_obj_id=selected_class_id
        ).values_list('subject', flat=True)
        faculty_marks = Marks.objects.filter(
            student__class_obj_id=selected_class_id,
            student__section__in=assignments.values('section'),
            subject__in=faculty_subjects,
            exam=latest_exam   
        )
        all_marks = Marks.objects.filter(
            student__class_obj_id=selected_class_id,
            student__section__in=assignments.values('section')
        )

        # 1. Class Average
        class_avg = faculty_marks.aggregate(avg=Avg('total_marks'))['avg'] or 0

        # 2. At-risk students (<50%)
        at_risk = faculty_marks.filter(total_marks__lt=50).select_related('student', 'subject')

        # 3. Subject-wise Average
        subject_avg = (
            all_marks.values('subject__subject_name')
            .annotate(avg=Avg('total_marks'))
        )
        top_student = faculty_marks.order_by('-total_marks').first()

        # Attendance percentage
        total_classes = Attendance.objects.filter(
            student__class_obj_id=selected_class_id
            ).count()

        present_count = Attendance.objects.filter(
            student__class_obj_id=selected_class_id,status='P').count()

        attendance_percent = (present_count / total_classes * 100) if total_classes > 0 else 0
        # -----------------------------
        # K-MEANS STUDENT CLUSTERING

        student_cluster_data = []

        class_students = Student.objects.filter(
            class_obj_id=selected_class_id
        )

        for student in class_students:
        # ONLY faculty handled subjects
            avg_marks = Marks.objects.filter(
             student=student,
             subject__in=faculty_subjects,
             exam=latest_exam
            ).aggregate(avg=Avg("total_marks"))["avg"] or 0

            total_days = Attendance.objects.filter(student=student).count()
            present_days = Attendance.objects.filter(
                student=student,
                status="P"
            ).count()

            attendance = (present_days / total_days * 100) if total_days > 0 else 0

            student_cluster_data.append({
                "student": student,
                "avg_marks": round(avg_marks, 2),
                "attendance": round(attendance, 2)
        })

        cluster_result = cluster_students(student_cluster_data)
        selected_cluster = request.GET.get("cluster")

        filtered_cluster_students = cluster_result["clusters"]

        if selected_cluster:
            filtered_cluster_students = [
                s for s in filtered_cluster_students
                if s["cluster"] == selected_cluster
            ]

        # ML Prediction
        predicted_pass_percent = predict_pass_percentage(class_avg, attendance_percent)

        total_students = Student.objects.filter(class_obj_id=selected_class_id).count()
        predicted_pass_students = int((predicted_pass_percent / 100) * total_students)


        data.update({
            "faculty": faculty,
            "class_avg": round(class_avg, 2),
            "at_risk": at_risk,
            "subject_avg": subject_avg,
            "top_student": top_student,
            "class_avg": round(class_avg, 2),
            "predicted_pass_percent": predicted_pass_percent,
            "predicted_pass_students": predicted_pass_students,
            "total_students": total_students,
            "is_class_incharge": is_class_incharge,
            "cluster_counts": cluster_result["counts"],
            "cluster_students": filtered_cluster_students,
            "selected_cluster": selected_cluster,
            "top_count": cluster_result["counts"].get("Top Performers", 0),
            "avg_count": cluster_result["counts"].get("Average Students", 0),
            "needs_count": cluster_result["counts"].get("Needs Attention", 0),
})

    return render(request, "faculty/performance.html", data)


from openpyxl import Workbook
from openpyxl.styles import Font,Alignment,PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils import timezone
import datetime
from django.contrib import messages
from django.contrib.auth.decorators import login_required


@login_required
def generate_attendance_report(request):
    faculty = Faculty.objects.get(user=request.user)

    class_incharge = FacultyAssignment.objects.filter(
        faculty=faculty,
        is_class_incharge=True
    ).first()

    if not class_incharge:
        return HttpResponse("Not authorized")

    class_obj = class_incharge.class_obj
    section = class_incharge.section
    from_date = request.POST.get("from_date")
    to_date = request.POST.get("to_date")

    if not from_date or not to_date:
       return HttpResponse("Please select both From and To dates")

    start_date = datetime.datetime.strptime(from_date, "%Y-%m-%d").date()
    end_date = datetime.datetime.strptime(to_date, "%Y-%m-%d").date()

# Safety check
    if start_date > end_date:
      return HttpResponse("From date cannot be greater than To date")

    students = Student.objects.filter(class_obj=class_obj, section=section)

    # ---- FIX: START DATE SHOULD NOT GO BEFORE FIRST ATTENDANCE ----
    first_record = Attendance.objects.filter(
        student__in=students
    ).order_by("date").first()

    if first_record and first_record.date > start_date:
        start_date = first_record.date

    # ---- CREATE DATE LIST ----
    date_list = []
    d = start_date
    while d <= end_date:
        date_list.append(d)
        d += datetime.timedelta(days=1)

    # ---- CREATE EXCEL ----
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # ===== STYLES =====
    title_font = Font(size=14, bold=True)
    header_font = Font(bold=True, color="FFFFFF")
    bold_font = Font(bold=True)

    center = Alignment(horizontal="center", vertical="center")

    yellow_fill = PatternFill("solid", start_color="D4A017")
    header_fill = PatternFill("solid", start_color="1F4E78")
    green_fill = PatternFill("solid", start_color="C6EFCE")
    red_fill = PatternFill("solid", start_color="FFC7CE")
    grey_fill = PatternFill("solid", start_color="E7E6E6")

    # ===== TITLE =====
    last_col = len(date_list) + 2
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws["A1"] = "ATTENDANCE REPORT"
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws["A1"].fill = yellow_fill

    # ===== CLASS INFO =====
    ws["A2"] = "Class"
    ws["B2"] = f"{class_obj.class_name} - {section.section_name}"

    ws["A3"] = "From"
    ws["B3"] = str(start_date)

    ws["A4"] = "To"
    ws["B4"] = str(end_date)

    for row in range(2, 5):
        ws[f"A{row}"].font = bold_font

    # ===== HEADER =====
    header_row = 6
    header = ["Roll No", "Name"] + [str(d) for d in date_list]
    ws.append(header)

    for col_num in range(1, len(header) + 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # ===== STUDENT DATA + DAILY COUNTS =====
    daily_present = {d: 0 for d in date_list}
    daily_absent = {d: 0 for d in date_list}

    for student in students:
        row = [student.roll_no, student.name]

        for d in date_list:
            if d.weekday() == 6:
                row.append("Holiday")
                continue

            record = Attendance.objects.filter(student=student, date=d).first()

            if record and record.status == "P":
                row.append("Present")
                daily_present[d] += 1
            elif record and record.status == "A":
                row.append("Absent")
                daily_absent[d] += 1
            else:
                row.append("")

        ws.append(row)

    # ===== TOTAL ROWS =====
    ws.append([])
    total_row = ["Total", ""]
    present_row = ["Present", ""]
    absent_row = ["Absent", ""]

    for d in date_list:
        total_row.append(daily_present[d] + daily_absent[d])
        present_row.append(daily_present[d])
        absent_row.append(daily_absent[d])

    ws.append(total_row)
    ws.append(present_row)
    ws.append(absent_row)

    # ===== STYLE TOTALS =====
    for row in range(ws.max_row - 2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = bold_font
            cell.alignment = center
            cell.fill = grey_fill

    # ===== CELL ALIGNMENT + COLORS =====
    for row in ws.iter_rows(min_row=7, max_row=ws.max_row - 4):
        for cell in row:
            cell.alignment = center

            if cell.value == "Present":
                cell.fill = green_fill
            elif cell.value == "Absent":
                cell.fill = red_fill
            elif cell.value == "Holiday":
                cell.fill = grey_fill

    # ===== AUTO WIDTH =====
    for column in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(column)

        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=column)

            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = max_length + 3

        ws.column_dimensions[column_letter].width = max_length + 2
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"attendance_{class_obj.class_name}_{section.section_name}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'

    wb.save(response)
    return response