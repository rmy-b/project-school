from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from details.models import Class, Section, Subject, Faculty, FacultyAssignment
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Alignment, Side

@login_required
def manage_classes(request):

    if request.method == "POST":

        action = request.POST.get("action")

        # ADD CLASS
        if action == "add_class":
            name = request.POST.get("class_name")
            if name:
                Class.objects.create(class_name=name)

        # MANAGE SECTIONS
        elif action == "manage_sections":
            class_id = request.POST.get("class_id")
            selected_sections = request.POST.getlist("sections")

            class_obj = Class.objects.get(id=class_id)

            existing_sections = Section.objects.filter(class_obj=class_obj)

            #DELETE unchecked
            for sec in existing_sections:
                if sec.section_name not in selected_sections:
                    sec.delete()

            # ADD checked
            for sec in selected_sections:
                if not Section.objects.filter(class_obj=class_obj, section_name=sec).exists():
                    Section.objects.create(class_obj=class_obj, section_name=sec)

        # MANAGE SUBJECTS
        elif action == "manage_subjects":
            class_id = request.POST.get("class_id")
            selected_subjects = request.POST.getlist("subjects")

            class_obj = Class.objects.get(id=class_id)

            existing_subjects = Subject.objects.filter(class_obj=class_obj)

            for sub in existing_subjects:
                if sub.subject_name not in selected_subjects:
                    sub.delete()

            for sub in selected_subjects:
                if not Subject.objects.filter(class_obj=class_obj, subject_name=sub).exists():
                    Subject.objects.create(class_obj=class_obj, subject_name=sub)

        return redirect("manage_classes")

    classes = Class.objects.all().prefetch_related('section_set', 'subject_set')

    return render(request, "adminpanel/manage_classes.html", {
        "classes": classes
    })

@login_required
def class_section_detail(request, class_id, section_id):

    class_obj = Class.objects.get(id=class_id)
    section = Section.objects.get(id=section_id)

    subjects = Subject.objects.filter(class_obj=class_obj)

    assignments = FacultyAssignment.objects.filter(
        class_obj=class_obj,
        section=section
    )

    # class incharge
    class_incharge = assignments.filter(is_class_incharge=True).first()

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "assign_faculty":
            faculty_id = request.POST.get("faculty_id")
            subject_id = request.POST.get("subject_id")
            is_incharge = request.POST.get("is_class_incharge") == "true"

            faculty = Faculty.objects.get(id=faculty_id)

            if is_incharge:

                # CHECK 1: Faculty must already handle a subject in this class+section
                existing = FacultyAssignment.objects.filter(
                    faculty=faculty,
                    class_obj=class_obj,
                    section=section
                ).first()

                if not existing:
                    request.session['error'] = f"{faculty.name} cannot be assigned as class in-charge because they are not handling any subject in Class {class_obj.class_name} Section {section.section_name}."
                    return redirect("class_section_detail", class_id=class_id, section_id=section_id)

                # CHECK 2: Faculty should NOT be incharge of another class
                already_incharge = FacultyAssignment.objects.filter(
                    faculty=faculty,
                    is_class_incharge=True
                ).exclude(
                    class_obj=class_obj,
                    section=section
                ).first()

                if already_incharge:
                    request.session['error'] = f"{faculty.name} is already assigned as class in-charge for Class {already_incharge.class_obj.class_name} Section {already_incharge.section.section_name}."
                    return redirect("class_section_detail", class_id=class_id, section_id=section_id)

                # NOW SAFE → update incharge

                # remove old incharge
                FacultyAssignment.objects.filter(
                    class_obj=class_obj,
                    section=section,
                    is_class_incharge=True
                ).update(is_class_incharge=False)

                # mark this faculty as incharge
                existing.is_class_incharge = True
                existing.save()

            else:
                subject = Subject.objects.get(id=subject_id)

                FacultyAssignment.objects.filter(
                    subject=subject,
                    class_obj=class_obj,
                    section=section
                ).delete()
                    
                FacultyAssignment.objects.filter(
                    faculty=faculty,
                    class_obj=class_obj,
                    section=section,
                    is_class_incharge=False
                ).delete()

                FacultyAssignment.objects.create(
                    faculty=faculty,
                    subject=subject,
                    class_obj=class_obj,
                    section=section,
                    is_class_incharge=False
                )

            return redirect("class_section_detail", class_id=class_id, section_id=section_id)

    error = request.session.pop('error',None)

    context = {
        "class_obj": class_obj,
        "section": section,
        "subjects": subjects,
        "assignments": assignments,
        "class_incharge": class_incharge,
        "faculty_list": Faculty.objects.all(),
        "error":error
    }

    return render(request, "adminpanel/class_section_detail.html", context)

def export_classes_excel(request):

    wb = Workbook()
    ws = wb.active
    ws.title = "Classes Report"

    # STYLES 
    header_font = Font(bold=True, size=13)
    sub_header_font = Font(bold=True, size=11)
    normal_font = Font(size=11)

    center_align = Alignment(vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # column width
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30

    row_num = 1

    classes = Class.objects.all()

    for class_obj in classes:
        sections = Section.objects.filter(class_obj=class_obj)

        for section in sections:

            assignments = FacultyAssignment.objects.filter(
                class_obj=class_obj,
                section=section
            )

            # INCHARGE 
            incharge = assignments.filter(is_class_incharge=True).first()
            incharge_name = incharge.faculty.name if incharge else "Not Assigned"

            # CLASS HEADER 
            ws.cell(row=row_num, column=1).value = f"Class {class_obj.class_name} - Section {section.section_name}"
            ws.cell(row=row_num, column=2).value = f"Incharge: {incharge_name}"

            ws.cell(row=row_num, column=1).font = header_font
            ws.cell(row=row_num, column=2).font = header_font

            row_num += 1

            # LINE BELOW HEADER 
            for col in range(1, 3):
                ws.cell(row=row_num, column=col).border = Border(bottom=Side(style='medium'))

            row_num += 1

            # TABLE HEADER
            ws.cell(row=row_num, column=1).value = "Subject"
            ws.cell(row=row_num, column=2).value = "Handled By"

            ws.cell(row=row_num, column=1).font = sub_header_font
            ws.cell(row=row_num, column=2).font = sub_header_font

            # LINE BELOW TABLE HEADER
            for col in range(1, 3):
                ws.cell(row=row_num, column=col).border = Border(bottom=Side(style='medium'))

            row_num += 1

            # SUBJECT ROWS
            subjects = Subject.objects.filter(class_obj=class_obj)

            for subject in subjects:
                faculty_name = ""

                for assign in assignments:
                    if assign.subject and assign.subject.id == subject.id:
                        faculty_name = assign.faculty.name

                if not faculty_name:
                    faculty_name = "Not Assigned"

                ws.cell(row=row_num, column=1).value = subject.subject_name
                ws.cell(row=row_num, column=2).value = faculty_name

                ws.cell(row=row_num, column=1).font = normal_font
                ws.cell(row=row_num, column=2).font = normal_font

                # BORDER FOR EACH ROW
                ws.cell(row=row_num, column=1).border = thin_border
                ws.cell(row=row_num, column=2).border = thin_border

                row_num += 1

            # SPACE BETWEEN SECTIONS 
            row_num += 2

    # DOWNLOAD 
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=classes_report.xlsx'

    wb.save(response)

    return response